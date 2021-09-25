
from openpyxl import Workbook
def createColumn(char, title):
    sheet[char+"1"] = title
    toDo = str(input('Give a toDo item: '))
    i = 2
    while str(toDo) != 'end':
        j = char+str(i)
        sheet[j] = str(toDo)
        toDo = str(input('Give a toDo item: '))
        if str(toDo) != 'end':
            i+=1
        else:
            break
    print(i)
def createRow(char, row, title):
    sheet['A'+str(row)] = title
    toDo = str(input('Give a toDo item: '))
    i = 1
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    while str(toDo) != 'end':
        j = letters[i]+str(row)
        sheet[j] = str(toDo)
        toDo = str(input('Give a toDo item: '))
        if str(toDo) != 'end':
            i+=1
        else:
            break
    print(letters[i])

workbook = Workbook()
sheet = workbook.active
createColumn('A', 'toDo')
createRow('A', 9, 'sum')
workbook.save(filename="toDo.xlsx")


'''
from openpyxl import load_workbook         # import load_workbook                  
filepath="/home/ubuntu/demo.xlsx"          # set file path          
wb=load_workbook(filepath)                 # load demo.xlsx     
sheet=wb.active                            # select demo.xlsx     
sheet['A1'] = 1                            # set value for cell A1=1      
sheet.cell(row=2, column=2).value = 2      # set value for cell B2=2  
wb.save(filepath)                          # save workbook    
'''
